"""Excel export using openpyxl.

Produces a three-sheet workbook:
  Sheet 1 — Master Timetable (days × periods grid)
  Sheet 2 — By Department   (sorted list)
  Sheet 3 — By Lecturer     (sorted list)
"""
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import Day, Period
from scheduling.models import TimetableEntry

# Design-system palette (hex without #)
_ACCENT      = "C7401F"
_ACCENT_TINT = "F9E8E3"
_RULE        = "D7D2C6"
_SURFACE     = "FBF9F4"
_INK         = "1E1E1C"
_INK_SOFT    = "4A4A46"
_MONO        = "Courier New"


def _border(color=_RULE):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _font(bold=False, color=_INK, size=9):
    return Font(bold=bold, color=color, size=size, name=_MONO)


def _header_font():
    return Font(bold=True, color=_INK, size=10, name=_MONO)


def _meta_row(ws, row_num, text, col_count):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    c = ws.cell(row=row_num, column=1, value=text)
    c.font = _font(color=_INK_SOFT, size=9)
    ws.row_dimensions[row_num].height = 15


def generate_excel(timetable) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    periods = list(Period.objects.order_by("index"))
    day_choices = list(Day.choices)
    days = [dv for dv, _ in day_choices]
    day_labels = [dl for _, dl in day_choices]
    day_order = {dv: i for i, dv in enumerate(days)}

    entries_all = list(
        TimetableEntry.objects.filter(timetable=timetable).select_related(
            "course__department", "lecturer", "venue", "start_period"
        )
    )

    # ------------------------------------------------------------------
    # Sheet 1: Master grid
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Master Timetable"
    col_count = len(days) + 1

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=timetable.name)
    title_cell.font = Font(bold=True, color=_INK, size=13, name="Calibri")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Metadata row
    gen_date = timetable.generated_at.strftime("%d %b %Y %H:%M") if timetable.generated_at else "—"
    meta = (
        f"Session: {timetable.academic_session}  |  "
        f"Status: {timetable.get_status_display()}  |  "
        f"Generated: {gen_date}  |  "
        f"Sessions: {timetable.scheduled_count}  |  "
        f"Soft score: {timetable.soft_score}"
    )
    _meta_row(ws, 2, meta, col_count)

    # Column headers (row 3)
    ws.row_dimensions[3].height = 18
    ws.column_dimensions["A"].width = 11

    hdr_period = ws.cell(row=3, column=1, value="Period")
    hdr_period.font = _header_font()
    hdr_period.fill = _fill(_SURFACE)
    hdr_period.border = _border()
    hdr_period.alignment = Alignment(horizontal="center", vertical="center")

    for ci, (dv, dl) in enumerate(day_choices, start=2):
        c = ws.cell(row=3, column=ci, value=dl)
        c.font = _header_font()
        c.fill = _fill(_SURFACE)
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = 24

    # Build entry map
    entry_map: dict = defaultdict(list)
    for e in entries_all:
        entry_map[(e.day, e.start_period.index)].append(e)

    # Grid rows
    for ri, period in enumerate(periods, start=4):
        ws.row_dimensions[ri].height = 72

        period_cell = ws.cell(
            row=ri, column=1,
            value=f"P{period.index}\n{period.start_time.strftime('%H:%M')}"
        )
        period_cell.font = _font(color=_INK_SOFT, size=8)
        period_cell.fill = _fill(_SURFACE)
        period_cell.border = _border()
        period_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for ci, dv in enumerate(days, start=2):
            cell_entries = entry_map.get((dv, period.index), [])
            if cell_entries:
                lines = []
                for e in cell_entries:
                    span = f" ×{e.length}p" if e.length > 1 else ""
                    lines += [f"{e.course.code}{span}", f"  {e.lecturer.name}", f"  {e.venue.name}", ""]
                xc = ws.cell(row=ri, column=ci, value="\n".join(lines).strip())
                xc.font = _font(size=8)
                xc.fill = _fill(_ACCENT_TINT)
            else:
                xc = ws.cell(row=ri, column=ci, value="")
                xc.font = _font(color=_RULE, size=8)
                xc.fill = _fill("FFFFFF")
            xc.border = _border()
            xc.alignment = Alignment(vertical="top", wrap_text=True)

    # ------------------------------------------------------------------
    # Sheet 2: By Department
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("By Department")
    sorted_by_dept = sorted(
        entries_all,
        key=lambda e: (e.course.department.code, day_order.get(e.day, 99), e.start_period.index)
    )
    _write_list_sheet(ws2, timetable, sorted_by_dept, group_label="Department",
                      group_fn=lambda e: e.course.department.name)

    # ------------------------------------------------------------------
    # Sheet 3: By Lecturer
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("By Lecturer")
    sorted_by_lect = sorted(
        entries_all,
        key=lambda e: (e.lecturer.name, day_order.get(e.day, 99), e.start_period.index)
    )
    _write_list_sheet(ws3, timetable, sorted_by_lect, group_label="Lecturer",
                      group_fn=lambda e: e.lecturer.name)

    return wb


def _write_list_sheet(ws, timetable, entries, group_label, group_fn):
    headers = [group_label, "Day", "Period", "Course", "Title", "Venue", "Size"]
    widths   = [22, 11, 9, 14, 30, 14, 7]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _header_font()
        c.fill = _fill(_SURFACE)
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 18

    prev_group = None
    for row, e in enumerate(entries, start=2):
        group = group_fn(e)
        row_data = [
            group if group != prev_group else "",
            e.get_day_display(),
            f"P{e.start_period.index}",
            e.course.code,
            e.course.title,
            e.venue.name,
            e.course.class_size,
        ]
        prev_group = group
        alt_fill = _fill("FAFAFA") if row % 2 == 0 else _fill("FFFFFF")
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = _font(size=9)
            c.fill = alt_fill
            c.border = _border()
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 16
